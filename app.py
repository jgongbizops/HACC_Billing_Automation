import warnings

warnings.filterwarnings(
    "ignore",
    message=r"Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
    module=r"openpyxl\..*",
)

import io
from datetime import date
from typing import List, Dict, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ---------------- CONFIG ----------------
MRC_RATE = 1.42

# Your file may have different column names month-to-month.
# We use "candidates" to auto-detect the right column even if it changes a bit.
CANDIDATES: Dict[str, List[str]] = {
    "iccid": ["ICCID", "ICCID Number", "SIM ICCID"],
    "vin": ["VIN", "Vehicle VIN"],
    "meid": ["MEID", "MEID/IMEI", "IMEI"],
    "mdn": ["MDN", "MDN/MSISDN", "MSISDN", "Phone Number"],
    "brand": ["BRAND", "Brand", "Service", "OEM"],
    "use_purpose": ["USE_PURPOSE", "Use Purpose", "USER PURPOSE", "USER_PURPOSE"],
    "device_group": ["DEVICE_GROUP_NAME", "Device Group", "Device Group Name"],

    # Usage files
    "roaming": ["Roaming Zone", "Roaming", "Roam", "Domestic/International"],
    "data_volume": ["Data Volume (MB)", "Data Volume", "Data Usage", "Usage (MB)", "Total Data (MB)"],
    "sms_volume": ["SMS Volume (msg)", "SMS Volume", "SMS Usage", "Messages", "Total Messages"],
    "voice_volume": ["Voice Volume", "Voice Usage", "Voice Minutes", "Minutes", "Total Minutes", "Usage (Min)", "Included Voice (m:ss)", "Voice Volume (m:ss)", "Voice", "Total Voice"],
}

# Sheet names in your hacc workbook (adjust if needed)
SHEETS = {
    "pre_rdr_tab": "Pre-RDR",
    "enrolled_ppu_tab": "Enrolled-PPU",
    "mrc_tab": "MRC(H,G)-Enrolled(5,10,30,50,1)",
}

# ---------------- HELPERS ----------------
def pick_first_existing_col(df: pd.DataFrame, candidates: List[str], label: str) -> str:
    cols = list(df.columns)
    colset = set(cols)
    for c in candidates:
        if c in colset:
            return c
    raise KeyError(f"Could not find {label} column. Available columns: {cols}")

def normalize_service_brand(series: pd.Series) -> pd.Series:
    # Rules you described:
    # - Genesis stays Genesis
    # - blanks -> Hyundai (H)
    # - H -> Hyundai, G -> Genesis
    s = series.fillna("H").astype(str).str.strip()
    up = s.str.upper()
    out = up.replace({
        "G": "Genesis",
        "GENESIS": "Genesis",
        "H": "Hyundai",
        "HYUNDAI": "Hyundai",
        "": "Hyundai",
    })
    # If it contains "GENESIS" in any form
    out = out.where(~up.str.contains("GENESIS", na=False), "Genesis")
    return out

def apply_common_device_filters(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    iccid_col = pick_first_existing_col(df, CANDIDATES["iccid"], "ICCID")
    df = df[df[iccid_col].notna()]

    # USE_PURPOSE in (1,2)
    if any(c in df.columns for c in CANDIDATES["use_purpose"]):
        use_purpose_col = pick_first_existing_col(df, CANDIDATES["use_purpose"], "USE_PURPOSE")
        df = df[df[use_purpose_col].isin([1, 2])]

    # Exclude test devices
    if any(c in df.columns for c in CANDIDATES["device_group"]):
        dg_col = pick_first_existing_col(df, CANDIDATES["device_group"], "DEVICE_GROUP_NAME")
        df = df[~df[dg_col].astype(str).str.upper().str.contains("TEST", na=False)]

    # Service mapping from BRAND (empty -> Hyundai)
    if any(c in df.columns for c in CANDIDATES["brand"]):
        brand_col = pick_first_existing_col(df, CANDIDATES["brand"], "BRAND")
        df["Service"] = normalize_service_brand(df[brand_col])
    else:
        df["Service"] = "Hyundai"

    # Standardize key columns
    df["__ICCID__"] = df[iccid_col].astype(str).str.strip()

    # Optional columns (safe if missing)
    for key, cand_list in [("VIN", CANDIDATES["vin"]), ("MEID", CANDIDATES["meid"]), ("MDN", CANDIDATES["mdn"])]:
        if any(c in df.columns for c in cand_list):
            df[f"__{key}__"] = df[pick_first_existing_col(df, cand_list, key)].astype(str).str.strip()
        else:
            df[f"__{key}__"] = ""

    return df

def roaming_flag(z) -> str:
    # Treat anything not clearly "roaming" as No
    s = str(z).strip().upper()
    if s in {"R", "ROAM", "ROAMING", "YES", "Y", "INTERNATIONAL"}:
        return "Yes"
    return "No"

def normalize_data_usage(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    iccid_col = pick_first_existing_col(df, CANDIDATES["iccid"], "ICCID (data usage)")
    roam_col = pick_first_existing_col(df, CANDIDATES["roaming"], "Roaming (data usage)")
    vol_col = pick_first_existing_col(df, CANDIDATES["data_volume"], "Data volume (MB)")

    df["__ICCID__"] = df[iccid_col].astype(str).str.strip()
    df["__ROAM__"] = df[roam_col].apply(roaming_flag)

    df["__VAL__"] = (
        df[vol_col].astype(str)
        .str.replace(",", "")
        .str.strip()
        .replace("", "0")
    )
    df["__VAL__"] = pd.to_numeric(df["__VAL__"], errors="coerce").fillna(0.0)
    return df[["__ICCID__", "__ROAM__", "__VAL__"]]

def normalize_sms_usage(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    iccid_col = pick_first_existing_col(df, CANDIDATES["iccid"], "ICCID (sms usage)")
    roam_col = pick_first_existing_col(df, CANDIDATES["roaming"], "Roaming (sms usage)")
    vol_col = pick_first_existing_col(df, CANDIDATES["sms_volume"], "SMS volume")

    df["__ICCID__"] = df[iccid_col].astype(str).str.strip()
    df["__ROAM__"] = df[roam_col].apply(roaming_flag)

    df["__VAL__"] = (
        df[vol_col].astype(str)
        .str.replace(",", "")
        .str.strip()
        .replace("", "0")
    )
    df["__VAL__"] = pd.to_numeric(df["__VAL__"], errors="coerce").fillna(0.0)
    return df[["__ICCID__", "__ROAM__", "__VAL__"]]

def normalize_voice_usage(df: pd.DataFrame) -> pd.DataFrame:
    """
    Voice can be either:
    - numeric minutes (e.g., 12)
    - m:ss format (e.g., 12:30)
    We'll convert to decimal minutes.
    """
    df = df.copy()
    iccid_col = pick_first_existing_col(df, CANDIDATES["iccid"], "ICCID (voice usage)")
    roam_col = pick_first_existing_col(df, CANDIDATES["roaming"], "Roaming (voice usage)")
    vol_col = pick_first_existing_col(df, CANDIDATES["voice_volume"], "Voice volume/minutes")

    df["__ICCID__"] = df[iccid_col].astype(str).str.strip()
    df["__ROAM__"] = df[roam_col].apply(roaming_flag)

    s = df[vol_col].astype(str).str.strip()

    has_colon = s.str.contains(":", na=False)
    minutes_plain = pd.to_numeric(s.where(~has_colon, ""), errors="coerce").fillna(0)

    parts = s.where(has_colon, "")
    mm = pd.to_numeric(parts.str.split(":", expand=True)[0], errors="coerce").fillna(0)
    ss = pd.to_numeric(parts.str.split(":", expand=True)[1], errors="coerce").fillna(0)
    minutes_colon = mm + (ss / 60.0)

    df["__VAL__"] = minutes_plain.where(~has_colon, minutes_colon).astype(float)
    return df[["__ICCID__", "__ROAM__", "__VAL__"]]

def inner_join_on_iccid(devices: pd.DataFrame, check_df: pd.DataFrame) -> pd.DataFrame:
    check = apply_common_device_filters(check_df)
    dev = devices.copy()
    dev = dev.merge(check[["__ICCID__"]].drop_duplicates(), on="__ICCID__", how="inner")
    return dev

def build_setup_devices(hacc_excel: pd.ExcelFile, pre_rdr_check_df: pd.DataFrame) -> pd.DataFrame:
    pre_rdr_hacc = pd.read_excel(hacc_excel, sheet_name=SHEETS["pre_rdr_tab"])
    pre_rdr_hacc = apply_common_device_filters(pre_rdr_hacc)

    # You said: if BRAND blank in hacc, still use device list from checks.
    # We filter by ICCIDs present in pre_rdr_check_df.
    return inner_join_on_iccid(pre_rdr_hacc, pre_rdr_check_df)

def build_ppu_devices(hacc_excel: pd.ExcelFile, enrolled_ppu_check_df: pd.DataFrame) -> pd.DataFrame:
    ppu_hacc = pd.read_excel(hacc_excel, sheet_name=SHEETS["enrolled_ppu_tab"])
    ppu_hacc = apply_common_device_filters(ppu_hacc)
    return inner_join_on_iccid(ppu_hacc, enrolled_ppu_check_df)

def build_mrc_devices(hacc_excel: pd.ExcelFile) -> pd.DataFrame:
    mrc_hacc = pd.read_excel(hacc_excel, sheet_name=SHEETS["mrc_tab"])
    mrc_hacc = apply_common_device_filters(mrc_hacc)
    return mrc_hacc

def build_base_rows(dev: pd.DataFrame, record_type: str, bill_start, bill_end, subscription_charges: float = 0.0) -> pd.DataFrame:
    if dev.empty:
        return pd.DataFrame()

    out = pd.DataFrame()
    out["ICCID"] = dev["__ICCID__"]
    out["VIN"] = dev["__VIN__"]
    out["MEID/IMEI"] = dev["__MEID__"]
    out["MDN/MSISDN"] = dev["__MDN__"]
    out["Service"] = dev["Service"]
    out["Record Type"] = record_type
    out["Bill Start Date"] = bill_start
    out["Bill End Date"] = bill_end

    out["Voice Roaming Zone"] = ""
    out["SMS Roaming Zone"] = ""
    out["Data Roaming Zone"] = ""

    out["Voice Usage (Min)"] = 0.0
    out["SMS Usage"] = 0.0
    out["Data Usage (MB)"] = 0.0

    out["Subscription Plan"] = ""
    out["Subscription Charges"] = subscription_charges
    return out

def build_usage_rows(dev: pd.DataFrame, usage_norm: pd.DataFrame, usage_type: str, record_type: str, bill_start, bill_end) -> pd.DataFrame:
    """
    usage_norm is standardized to columns: __ICCID__, __ROAM__ ("Yes"/"No"), __VAL__
    Only devices found in usage file are included (inner join) - as you requested.
    """
    if dev.empty or usage_norm.empty:
        return pd.DataFrame()

    agg = usage_norm.groupby(["__ICCID__", "__ROAM__"], as_index=False)["__VAL__"].sum()

    merged = dev.merge(agg, on="__ICCID__", how="inner")

    rows = []
    for flag in ["No", "Yes"]:  # No first, then Yes
        sub = merged[merged["__ROAM__"] == flag]
        if sub.empty:
            continue

        base = build_base_rows(sub, record_type=record_type, bill_start=bill_start, bill_end=bill_end, subscription_charges=0.0)

        if usage_type == "data":
            base["Data Roaming Zone"] = flag
            base["Data Usage (MB)"] = sub["__VAL__"].values
        elif usage_type == "sms":
            base["SMS Roaming Zone"] = flag
            base["SMS Usage"] = sub["__VAL__"].values
        elif usage_type == "voice":
            base["Voice Roaming Zone"] = flag
            base["Voice Usage (Min)"] = sub["__VAL__"].values

        rows.append(base)

    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()

def build_invoice_summary_from_bill_detail(bill_detail_df: pd.DataFrame) -> pd.DataFrame:
    # Lightweight internal summary preview (your actual invoice output is the template-based file)
    if bill_detail_df.empty:
        return pd.DataFrame()

    df = bill_detail_df.copy()
    summary = pd.DataFrame({
        "Metric": [
            "Setup lines",
            "PPU subscription lines",
            "MRC lines",
            "Total Data MB",
            "Total SMS",
            "Total Voice Min",
        ],
        "Value": [
            df[df["Record Type"] == "TMS Service Setup"]["ICCID"].nunique(),
            df[df["Record Type"] == "TMS Service Enrolled - Subscription"]["ICCID"].nunique(),
            df[df["Record Type"] == "TMS Service Enrolled - MRC"]["ICCID"].nunique(),
            df["Data Usage (MB)"].sum(),
            df["SMS Usage"].sum(),
            df["Voice Usage (Min)"].sum(),
        ]
    })
    return summary

def process_hacc_billing(
    hacc_file,
    pre_rdr_check_file,
    enrolled_ppu_check_file,
    data_usage_file,
    sms_usage_file,
    voice_usage_file,
    bill_start_date,
    bill_end_date,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    hacc_excel = pd.ExcelFile(hacc_file)
    pre_rdr_check = pd.read_excel(pre_rdr_check_file)
    enrolled_ppu_check = pd.read_excel(enrolled_ppu_check_file)

    setup_devices = build_setup_devices(hacc_excel, pre_rdr_check)
    ppu_devices = build_ppu_devices(hacc_excel, enrolled_ppu_check)
    mrc_devices = build_mrc_devices(hacc_excel)

    data_usage_raw = pd.read_excel(data_usage_file)
    sms_usage_raw = pd.read_excel(sms_usage_file)
    voice_usage_raw = pd.read_excel(voice_usage_file)

    data_usage = normalize_data_usage(data_usage_raw)
    sms_usage = normalize_sms_usage(sms_usage_raw)
    voice_usage = normalize_voice_usage(voice_usage_raw)

    # 1) TMS Service Setup base rows
    setup_rows = build_base_rows(setup_devices, "TMS Service Setup", bill_start_date, bill_end_date, subscription_charges=0.0)

    # 2) Setup usage breakdown
    setup_data = build_usage_rows(setup_devices, data_usage, "data", "TMS Service Setup - Usage Breakdown", bill_start_date, bill_end_date)
    setup_sms = build_usage_rows(setup_devices, sms_usage, "sms", "TMS Service Setup - Usage Breakdown", bill_start_date, bill_end_date)
    setup_voice = build_usage_rows(setup_devices, voice_usage, "voice", "TMS Service Setup - Usage Breakdown", bill_start_date, bill_end_date)

    # 3) PPU subscription base rows (no usage yet)
    ppu_sub = build_base_rows(ppu_devices, "TMS Service Enrolled - Subscription", bill_start_date, bill_end_date, subscription_charges=0.0)

    # 4) PPU usage breakdown
    ppu_data = build_usage_rows(ppu_devices, data_usage, "data", "TMS Service Enrolled (PPU) - Usage breakdown", bill_start_date, bill_end_date)
    ppu_sms = build_usage_rows(ppu_devices, sms_usage, "sms", "TMS Service Enrolled (PPU) - Usage breakdown", bill_start_date, bill_end_date)
    ppu_voice = build_usage_rows(ppu_devices, voice_usage, "voice", "TMS Service Enrolled (PPU) - Usage breakdown", bill_start_date, bill_end_date)

    # 5) MRC rows @ 1.42
    mrc_rows = build_base_rows(mrc_devices, "TMS Service Enrolled - MRC", bill_start_date, bill_end_date, subscription_charges=MRC_RATE)

    bill_detail = pd.concat(
        [setup_rows, setup_data, setup_sms, setup_voice, ppu_sub, ppu_data, ppu_sms, ppu_voice, mrc_rows],
        ignore_index=True
    )

    invoice_summary_preview = build_invoice_summary_from_bill_detail(bill_detail)
    return bill_detail, invoice_summary_preview

# ---------------- INVOICE TEMPLATE WRITER ----------------
def write_invoice_from_bill_detail_to_template(
    bill_detail_df: pd.DataFrame,
    invoice_template_file,
    summary_sheet_name: str = "Summary",
) -> io.BytesIO:
    """
    Writes ONLY the input cells (counts + usage) into the existing invoice template Summary sheet.
    Leaves all Excel formulas intact.
    """

    df = bill_detail_df.copy()
    df["Service"] = df.get("Service", "Hyundai").fillna("Hyundai").astype(str).str.strip().replace({"H": "Hyundai", "G": "Genesis"})

    def uniq_lines(sub: pd.DataFrame) -> int:
        return int(sub["ICCID"].nunique()) if not sub.empty else 0

    def sum_num(sub: pd.DataFrame, col: str) -> float:
        if sub.empty or col not in sub.columns:
            return 0.0
        return float(pd.to_numeric(sub[col], errors="coerce").fillna(0).sum())

    setup_base = df[df["Record Type"] == "TMS Service Setup"]
    mrc_base = df[df["Record Type"] == "TMS Service Enrolled - MRC"]
    ppu_base = df[df["Record Type"] == "TMS Service Enrolled - Subscription"]

    setup_usage = df[df["Record Type"] == "TMS Service Setup - Usage Breakdown"]
    ppu_usage = df[df["Record Type"] == "TMS Service Enrolled (PPU) - Usage breakdown"]

    def usage(service: str, base: pd.DataFrame, col: str, roam_col: str, flag: str) -> float:
        sub = base[(base["Service"] == service) & (base[roam_col] == flag)]
        return sum_num(sub, col)

    wb = load_workbook(invoice_template_file)
    ws = wb[summary_sheet_name]

    def setc(cell: str, val):
        ws[cell].value = val

    # ---------------- HYUNDAI ----------------
    h_setup = uniq_lines(setup_base[setup_base["Service"] == "Hyundai"])
    for c in ["E4", "E5", "E6"]:
        setc(c, h_setup)

    setc("H4", usage("Hyundai", setup_usage, "Data Usage (MB)", "Data Roaming Zone", "No"))
    setc("K4", usage("Hyundai", setup_usage, "Data Usage (MB)", "Data Roaming Zone", "Yes"))
    setc("H5", usage("Hyundai", setup_usage, "SMS Usage", "SMS Roaming Zone", "No"))
    setc("K5", usage("Hyundai", setup_usage, "SMS Usage", "SMS Roaming Zone", "Yes"))
    setc("H6", usage("Hyundai", setup_usage, "Voice Usage (Min)", "Voice Roaming Zone", "No"))
    setc("K6", usage("Hyundai", setup_usage, "Voice Usage (Min)", "Voice Roaming Zone", "Yes"))

    setc("E7", uniq_lines(mrc_base[mrc_base["Service"] == "Hyundai"]))

    h_ppu = uniq_lines(ppu_base[ppu_base["Service"] == "Hyundai"])
    for c in ["E8", "E9", "E10"]:
        setc(c, h_ppu)

    setc("H8", usage("Hyundai", ppu_usage, "Data Usage (MB)", "Data Roaming Zone", "No"))
    setc("K8", usage("Hyundai", ppu_usage, "Data Usage (MB)", "Data Roaming Zone", "Yes"))
    setc("H9", usage("Hyundai", ppu_usage, "SMS Usage", "SMS Roaming Zone", "No"))
    setc("K9", usage("Hyundai", ppu_usage, "SMS Usage", "SMS Roaming Zone", "Yes"))
    setc("H10", usage("Hyundai", ppu_usage, "Voice Usage (Min)", "Voice Roaming Zone", "No"))
    setc("K10", usage("Hyundai", ppu_usage, "Voice Usage (Min)", "Voice Roaming Zone", "Yes"))

    # ---------------- GENESIS ----------------
    g_setup = uniq_lines(setup_base[setup_base["Service"] == "Genesis"])
    for c in ["E14", "E15", "E16"]:
        setc(c, g_setup)

    setc("H14", usage("Genesis", setup_usage, "Data Usage (MB)", "Data Roaming Zone", "No"))
    setc("K14", usage("Genesis", setup_usage, "Data Usage (MB)", "Data Roaming Zone", "Yes"))
    setc("H15", usage("Genesis", setup_usage, "SMS Usage", "SMS Roaming Zone", "No"))
    setc("K15", usage("Genesis", setup_usage, "SMS Usage", "SMS Roaming Zone", "Yes"))
    setc("H16", usage("Genesis", setup_usage, "Voice Usage (Min)", "Voice Roaming Zone", "No"))
    setc("K16", usage("Genesis", setup_usage, "Voice Usage (Min)", "Voice Roaming Zone", "Yes"))

    setc("E17", uniq_lines(mrc_base[mrc_base["Service"] == "Genesis"]))

    g_ppu = uniq_lines(ppu_base[ppu_base["Service"] == "Genesis"])
    for c in ["E18", "E19", "E20"]:
        setc(c, g_ppu)

    setc("H18", usage("Genesis", ppu_usage, "Data Usage (MB)", "Data Roaming Zone", "No"))
    setc("K18", usage("Genesis", ppu_usage, "Data Usage (MB)", "Data Roaming Zone", "Yes"))
    setc("H19", usage("Genesis", ppu_usage, "SMS Usage", "SMS Roaming Zone", "No"))
    setc("K19", usage("Genesis", ppu_usage, "SMS Usage", "SMS Roaming Zone", "Yes"))
    setc("H20", usage("Genesis", ppu_usage, "Voice Usage (Min)", "Voice Roaming Zone", "No"))
    setc("K20", usage("Genesis", ppu_usage, "Voice Usage (Min)", "Voice Roaming Zone", "Yes"))

    # Ensure Excel recalculates formulas on open
    wb.calculation.fullCalcOnLoad = True

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

from openpyxl import load_workbook

BILL_DETAIL_COLS = [
    "ID",
    "Invoice Number",
    "Bill Cycle Month",
    "ICCID",
    "VIN",
    "MEID/IMEI",
    "MDN/MSISDN",
    "Service",
    "Record Type",
    "Voice Roaming Zone",
    "SMS Roaming Zone",
    "Data Roaming Zone",
    "Bill Start Date",
    "Bill End Date",
    "Voice Usage (Min)",
    "SMS Usage",
    "Data Usage (MB)",
    "Subscription Plan",
    "Subscription Charges",
]

def write_bill_detail_to_template(bill_detail_df, bill_detail_template_file,
                                 sheet1="Detail Bill", sheet2="Detail Bill 2",
                                 start_row=2, max_rows_per_sheet=1_000_000):
    """
    Fills your Bill Detail TEMPLATE in the exact formatting:
    - keeps headers, widths, styles
    - clears old data from row 2+
    - writes new rows copying the style from the template's first data row
    - spills into 'Detail Bill 2' automatically
    Returns BytesIO for download.
    """
    # --- Load template workbook ---
    wb = load_workbook(bill_detail_template_file)
    if sheet1 not in wb.sheetnames:
        raise KeyError(f"Template is missing sheet: {sheet1}")
    ws1 = wb[sheet1]

    # Ensure sheet2 exists
    if sheet2 not in wb.sheetnames:
        wb.create_sheet(sheet2)
    ws2 = wb[sheet2]

    # --- Capture style row BEFORE clearing (row 2 in template) ---
    def capture_row_style(ws, row_idx):
        styles = []
        for col_idx in range(1, len(BILL_DETAIL_COLS) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            styles.append({
                "font": c.font,
                "fill": c.fill,
                "border": c.border,
                "alignment": c.alignment,
                "number_format": c.number_format,
                "protection": c.protection,
                "comment": c.comment,
            })
        return styles

    style1 = capture_row_style(ws1, start_row)

    # For ws2: if it's empty, copy header row + style row from ws1
    if ws2.max_row < 2 or ws2.cell(1, 1).value is None:
        # Copy header row values + styles
        for col_idx in range(1, len(BILL_DETAIL_COLS) + 1):
            src = ws1.cell(row=1, column=col_idx)
            dst = ws2.cell(row=1, column=col_idx, value=src.value)
            dst.font = src.font
            dst.fill = src.fill
            dst.border = src.border
            dst.alignment = src.alignment
            dst.number_format = src.number_format
            dst.protection = src.protection

        # Copy column widths
        for col_letter, dim in ws1.column_dimensions.items():
            ws2.column_dimensions[col_letter].width = dim.width

        # Make sure style row exists to copy formatting
        for col_idx in range(1, len(BILL_DETAIL_COLS) + 1):
            dst = ws2.cell(row=start_row, column=col_idx)
            s = style1[col_idx - 1]
            dst.font = s["font"]
            dst.fill = s["fill"]
            dst.border = s["border"]
            dst.alignment = s["alignment"]
            dst.number_format = s["number_format"]
            dst.protection = s["protection"]

    style2 = capture_row_style(ws2, start_row)

    # --- Clear old data (keep header row 1) ---
    def clear_data(ws):
        if ws.max_row >= start_row:
            ws.delete_rows(start_row, ws.max_row - start_row + 1)
        # recreate style row so we can copy formatting when writing
        for col_idx in range(1, len(BILL_DETAIL_COLS) + 1):
            ws.cell(row=start_row, column=col_idx)

    clear_data(ws1)
    clear_data(ws2)

    # restore style rows
    for col_idx in range(1, len(BILL_DETAIL_COLS) + 1):
        c1 = ws1.cell(row=start_row, column=col_idx)
        s1 = style1[col_idx - 1]
        c1.font = s1["font"]
        c1.fill = s1["fill"]
        c1.border = s1["border"]
        c1.alignment = s1["alignment"]
        c1.number_format = s1["number_format"]
        c1.protection = s1["protection"]

        c2 = ws2.cell(row=start_row, column=col_idx)
        s2 = style2[col_idx - 1]
        c2.font = s2["font"]
        c2.fill = s2["fill"]
        c2.border = s2["border"]
        c2.alignment = s2["alignment"]
        c2.number_format = s2["number_format"]
        c2.protection = s2["protection"]

    # --- Ensure dataframe has all required columns in correct order ---
    df = bill_detail_df.copy()
    for col in BILL_DETAIL_COLS:
        if col not in df.columns:
            df[col] = ""  # safe default

    df = df[BILL_DETAIL_COLS]

    # --- Write rows, spilling into sheet2 ---
    current_ws = ws1
    current_style = style1
    row_in_sheet = start_row
    global_id = 1

    for _, r in df.iterrows():
        # decide which sheet
        if (row_in_sheet - start_row) >= max_rows_per_sheet:
            current_ws = ws2
            current_style = style2
            row_in_sheet = start_row

        # write row
        for col_idx, col_name in enumerate(BILL_DETAIL_COLS, start=1):
            cell = current_ws.cell(row=row_in_sheet, column=col_idx)

            # apply style
            s = current_style[col_idx - 1]
            cell.font = s["font"]
            cell.fill = s["fill"]
            cell.border = s["border"]
            cell.alignment = s["alignment"]
            cell.number_format = s["number_format"]
            cell.protection = s["protection"]

            # set value (special for ID)
            if col_name == "ID":
                cell.value = global_id
            else:
                v = r[col_name]
                # Convert numpy to python scalars
                if hasattr(v, "item"):
                    v = v.item()
                cell.value = v

        row_in_sheet += 1
        global_id += 1

    # Force Excel to recalc on open (keeps formulas in template happy)
    wb.calculation.fullCalcOnLoad = True

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out



# ---------------- STREAMLIT UI ----------------
st.title("HACC Billing Automation (Bill Detail + Invoice Template Output)")

st.write("Upload files → generate Bill Detail + an Invoice file (template Summary filled).")

hacc_file = st.file_uploader("hacc_Sep25.xlsx", type=["xlsx"])
pre_rdr_check_file = st.file_uploader("pre-rdr_check_Sep25.xlsx", type=["xlsx"])
enrolled_ppu_check_file = st.file_uploader("enrolled-ppu_check_Sep25.xlsx", type=["xlsx"])
data_usage_file = st.file_uploader("HAC_Data.xlsx", type=["xlsx"])
sms_usage_file = st.file_uploader("HAC_SMS.xlsx", type=["xlsx"])
voice_usage_file = st.file_uploader("HAC_Voice.xlsx", type=["xlsx"])

invoice_template = st.file_uploader("Invoice Template (xlsx) - Summary tab will be filled", type=["xlsx"])
bill_detail_template = st.file_uploader("Bill Detail Template (xlsx)", type=["xlsx"])


col1, col2 = st.columns(2)
with col1:
    bill_start = st.date_input("Bill Start Date", value=date(2025, 9, 1))
with col2:
    bill_end = st.date_input("Bill End Date", value=date(2025, 9, 30))

if st.button("Generate Bill Detail + Invoice"):

    missing = [
        name for name, f in [
            ("hacc_Sep25", hacc_file),
            ("pre-rdr_check_Sep25", pre_rdr_check_file),
            ("enrolled-ppu_check_Sep25", enrolled_ppu_check_file),
            ("HAC_Data", data_usage_file),
            ("HAC_SMS", sms_usage_file),
            ("HAC_Voice", voice_usage_file),
        ]
        if f is None
    ]

    if missing:
        st.error(f"Please upload: {', '.join(missing)}")
        st.stop()

    # 1) create data
    bill_detail_df, invoice_summary_df = process_hacc_billing(
        hacc_file,
        pre_rdr_check_file,
        enrolled_ppu_check_file,
        data_usage_file,
        sms_usage_file,
        voice_usage_file,
        bill_start,
        bill_end,
    )

    # 2) Bill Detail TEMPLATE output
    if bill_detail_template is not None:
        bill_detail_df["Service"] = bill_detail_df["Service"].replace({
            "Hyundai": "H",
            "Genesis": "G",
        }).fillna("H")

        bill_detail_bytes = write_bill_detail_to_template(
            bill_detail_df=bill_detail_df,
            bill_detail_template_file=bill_detail_template,
            sheet1="Detail Bill",
            sheet2="Detail Bill 2",
            start_row=2,
        )

        st.download_button(
            label="Download Bill Detail Output (Template formatted)",
            data=bill_detail_bytes,
            file_name="Bill_Detail_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.warning("Upload the Bill Detail Template to generate the formatted Bill Detail output.")

    # 3) Invoice TEMPLATE output
    if invoice_template is not None:
        invoice_bytes = write_invoice_from_bill_detail_to_template(
            bill_detail_df=bill_detail_df,
            invoice_template_file=invoice_template,
            summary_sheet_name="Summary",
        )

        st.download_button(
            label="Download Invoice Output (Summary filled)",
            data=invoice_bytes,
            file_name="Invoice_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.warning("Upload the Invoice Template to generate the Invoice output.")

    # 4) UI feedback
    st.success("Processing complete!")

    st.subheader("Preview: Bill Detail (first 20 rows)")
    st.dataframe(bill_detail_df.head(20))

    st.subheader("Preview: Invoice Summary")
    st.dataframe(invoice_summary_df)
